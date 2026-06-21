"""把当前的报价 / 排期状态导出成 Excel（用户点"下载"时实时生成，反映看板当前状态）。"""
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import models
import quote_service

_HEAD_FILL = PatternFill("solid", fgColor="2563EB")
_HEAD_FONT = Font(color="FFFFFF", bold=True)
_PHASE_FILL = PatternFill("solid", fgColor="EEF2FF")
_BOLD = Font(bold=True)
_thin = Side(style="thin", color="E5E7EB")
_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = _HEAD_FILL
        cell.font = _HEAD_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


_PHASE_NAMES = {"A": "前期筹备", "B": "拍摄执行", "C": "后期制作", "D": "其他杂费"}


def build_quote_xlsx(db, project_id: str, version: str = "client") -> bytes:
    """version='client' 客户版(利润已摊进单价、不显示成本/利润)；'internal' 内部版(成本+利润+实收)。"""
    project = db.query(models.Project).filter(models.Project.id == project_id).first()
    q = quote_service.serialize_quote(db, project_id)
    items = q.get("items", [])
    totals = q.get("totals", {})
    internal = (version == "internal")

    wb = Workbook()
    ws = wb.active
    ws.title = "报价单(内部版)" if internal else "报价单"

    ws["A1"] = f"{project.name} · 报价单（{'内部版' if internal else '客户版'}）"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"{project.film_type or ''} · 约 {project.duration_minutes or 0} 分钟 · 拍摄 {project.shoot_days or 0} 天"
    ws["A2"].font = Font(color="6B7280")

    if internal:
        headers = ["阶段", "项目", "成本单价", "人数", "天数/数量", "单位", "成本金额", "客户单价", "客户金额"]
    else:
        headers = ["阶段", "项目", "单价", "人数", "天数/数量", "单位", "金额"]
    ncol = len(headers)
    hr = 4
    for i, h in enumerate(headers, 1):
        ws.cell(row=hr, column=i, value=h)
    _style_header(ws, hr, ncol)

    r = hr + 1
    for ph in ["A", "B", "C", "D"]:
        rows = [it for it in items if it["phase"] == ph]
        if not rows:
            continue
        ws.cell(row=r, column=1, value=f"{ph} · {_PHASE_NAMES[ph]}").font = _BOLD
        for c in range(1, ncol + 1):
            ws.cell(row=r, column=c).fill = _PHASE_FILL
        r += 1
        for it in rows:
            ws.cell(row=r, column=1, value=ph)
            ws.cell(row=r, column=2, value=it["item_name"])
            if internal:
                ws.cell(row=r, column=3, value=it["unit_price"])
                ws.cell(row=r, column=4, value=it["qty_people"])
                ws.cell(row=r, column=5, value=it["qty_days"])
                ws.cell(row=r, column=6, value=it["unit"])
                ws.cell(row=r, column=7, value=it["amount"])
                ws.cell(row=r, column=8, value=it["client_unit_price"])
                ws.cell(row=r, column=9, value=it["client_amount"])
            else:
                ws.cell(row=r, column=3, value=it["client_unit_price"])
                ws.cell(row=r, column=4, value=it["qty_people"])
                ws.cell(row=r, column=5, value=it["qty_days"])
                ws.cell(row=r, column=6, value=it["unit"])
                ws.cell(row=r, column=7, value=it["client_amount"])
            r += 1
        cost_sub = totals.get("subtotals", {}).get(ph, 0)
        client_sub = totals.get("client_subtotals", {}).get(ph, 0)
        ws.cell(row=r, column=2, value="小计").font = _BOLD
        if internal:
            ws.cell(row=r, column=7, value=cost_sub).font = _BOLD
            ws.cell(row=r, column=9, value=client_sub).font = _BOLD
        else:
            ws.cell(row=r, column=7, value=client_sub).font = _BOLD
        r += 1

    r += 1
    last_col = ncol
    if internal:
        summary = [
            ("成本核算（明细合计）", totals.get("cost_total", 0)),
            (f"利润（{round(totals.get('margin_rate', 0.25) * 100)}%）", totals.get("profit", 0)),
            ("对客户实收（含税）", totals.get("client_price", 0)),
        ]
    else:
        summary = [("合计（含税）", totals.get("client_price", 0))]
    for label, val in summary:
        ws.cell(row=r, column=last_col - 1, value=label).font = _BOLD
        ws.cell(row=r, column=last_col, value=val).font = _BOLD
        r += 1
    ws.cell(row=r - 1, column=last_col - 1).fill = _HEAD_FILL
    ws.cell(row=r - 1, column=last_col - 1).font = _HEAD_FONT
    ws.cell(row=r - 1, column=last_col).fill = _HEAD_FILL
    ws.cell(row=r - 1, column=last_col).font = _HEAD_FONT

    widths = [10, 18, 11, 8, 11, 8, 13, 11, 13][:ncol]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_schedule_xlsx(db, project_id: str) -> bytes:
    project = db.query(models.Project).filter(models.Project.id == project_id).first()
    s = quote_service.serialize_schedule(db, project_id)
    items = s.get("items", [])

    wb = Workbook()
    ws = wb.active
    ws.title = "执行排期"

    ws["A1"] = f"{project.name} · 执行排期"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"交付日 {project.delivery_date}（倒推） · 拍摄 {project.shoot_days or 0} 天"
    ws["A2"].font = Font(color="6B7280")

    headers = ["阶段", "任务", "开始", "结束", "关键节点", "需客户配合", "状态"]
    hr = 4
    for i, h in enumerate(headers, 1):
        ws.cell(row=hr, column=i, value=h)
    _style_header(ws, hr, len(headers))

    status_cn = {"completed": "已完成", "current": "进行中", "pending": "待办"}
    r = hr + 1
    for it in items:
        ws.cell(row=r, column=1, value=it["stage"])
        ws.cell(row=r, column=2, value=it["task"])
        ws.cell(row=r, column=3, value=it["start_date"])
        ws.cell(row=r, column=4, value=it["end_date"])
        ws.cell(row=r, column=5, value="★ 是" if it["is_milestone"] else "")
        ws.cell(row=r, column=6, value="是" if it["needs_client"] else "")
        ws.cell(row=r, column=7, value=status_cn.get(it["status"], it["status"]))
        r += 1

    widths = [8, 28, 12, 12, 10, 10, 8]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
