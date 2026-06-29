# 6.26 迭代全面测试（端到端，打 API）。运行：./venv/bin/python _fulltest.py
import urllib.request, json, uuid, io, time

BASE = "http://127.0.0.1:8000"
P, F = 0, 0
LINES = []

def log(s=""):
    print(s); LINES.append(s)

def call(method, path, body=None):
    data = json.dumps(body if body is not None else {}).encode() if method in ("POST", "PUT", "DELETE") else None
    r = urllib.request.Request(BASE + path, data=data, headers={"Content-Type": "application/json"}, method=method)
    return json.load(urllib.request.urlopen(r, timeout=90))

def raw(path):
    return urllib.request.urlopen(BASE + path, timeout=60).read()

def chat(msg, role, pid="p1"):
    r = urllib.request.Request(BASE + "/api/chat", data=json.dumps({"message": msg, "user_id": role, "role": role, "project_id": pid, "session_id": "default"}).encode(), headers={"Content-Type": "application/json"}, method="POST")
    return json.load(urllib.request.urlopen(r, timeout=90))["reply"]

def upload(path, fn, content):
    b = "----" + uuid.uuid4().hex
    body = (f'--{b}\r\nContent-Disposition: form-data; name="file"; filename="{fn}"\r\n\r\n').encode() + content + f'\r\n--{b}--\r\n'.encode()
    r = urllib.request.Request(BASE + path, data=body, headers={"Content-Type": f"multipart/form-data; boundary={b}"}, method="POST")
    return json.load(urllib.request.urlopen(r, timeout=60))

def check(name, cond, detail=""):
    global P, F
    if cond: P += 1; log(f"  ✅ {name}" + (f"  [{detail}]" if detail else ""))
    else:    F += 1; log(f"  ❌ {name}  [{detail}]")

def find(items, name):
    return next((i for i in items if i["item_name"] == name), None)

log("="*72); log("AI 制片 · 6.26 迭代全面测试报告"); log("时间: " + time.strftime("%Y-%m-%d %H:%M:%S")); log("="*72)
h = call("GET", "/api/health"); log(f"模型: {h['model']} | key: {h['api_key_masked']}\n")

# ---------- F1 拍摄天数不被覆盖 ----------
log("【F1】新建项目拍摄天数不被生成覆盖")
c = call("POST", "/api/projects", {"name": "TEST-F1", "client": "测试", "film_type": "宣传片", "shoot_days": 4, "duration_minutes": 5})
fp = c["id"]
call("PUT", f"/api/projects/{fp}/brief", {"brief_text": "一支单天就能拍完的短视频，1天拍摄。"})
call("POST", f"/api/projects/{fp}/generate", {})
pj = call("GET", f"/api/projects/{fp}")
check("用户设4天, Brief说1天 → 仍为4天", pj["shoot_days"] == 4, f"shoot_days={pj['shoot_days']}")
sd = call("GET", f"/api/projects/{fp}/schedule")
shoot = [i for i in sd["items"] if "拍摄 DAY" in i["task"]]
check("排期拍摄节点按4天", shoot and "DAY1-4" in shoot[0]["task"], shoot[0]["task"] if shoot else "无")
call("DELETE", f"/api/projects/{fp}")

# ---------- F2 人天联动 ----------
log("\n【F2】改拍摄天数，人天项(含D段餐食/设备车)联动")
call("POST", "/api/projects/p1/generate", {})
chat("拍摄改成5天", "boss")
q = call("GET", "/api/projects/p1/quote")
ok = all((find(q["items"], n) or {}).get("qty_days") == 5 for n in ["导演", "设备车", "机动费用"])
check("导演/设备车/机动费用 天数全部=5", ok, "导演%s 设备车%s 机动%s" % (find(q['items'],'导演')['qty_days'], find(q['items'],'设备车')['qty_days'], find(q['items'],'机动费用')['qty_days']))

# ---------- B 报价可管理 ----------
log("\n【B】报价模块：增删/客户单价/锁定/反推")
call("POST", "/api/projects/p1/generate", {})
q = call("GET", "/api/projects/p1/quote"); dao = find(q["items"], "导演")
check("B4 字段齐全(成本/客户单价/锁定/毛利率)", all(k in dao for k in ("unit_price", "client_unit_price", "is_locked", "gross_margin")))
# B2 改客户单价独立于成本
call("PUT", f"/api/projects/p1/quote/items/{dao['id']}", {"client_unit_price": 8000})
q = call("GET", "/api/projects/p1/quote"); dao = find(q["items"], "导演")
check("B2 改客户单价→客户金额变、成本不变", dao["client_unit_price"] == 8000 and dao["unit_price"] == 3000, f"客户单价{dao['client_unit_price']} 成本{dao['unit_price']}")
# B3 锁定 + 批量利润率跳过锁定项
call("PUT", f"/api/projects/p1/quote/items/{dao['id']}", {"is_locked": True})
call("PUT", "/api/projects/p1/margin", {"margin_rate": 0.5})
q = call("GET", "/api/projects/p1/quote"); dao = find(q["items"], "导演"); sy = find(q["items"], "摄影")
check("B3 锁定项不被批量利润率改动", dao["client_unit_price"] == 8000, f"导演{dao['client_unit_price']}")
check("B3 未锁定项跟随利润率(摄影=2500×1.5)", sy["client_unit_price"] == 3750, f"摄影{sy['client_unit_price']}")
# B1 增删
n0 = len(q["items"])
r = call("POST", "/api/projects/p1/quote/items", {"phase": "D", "item_name": "场地租赁", "unit_price": 2000, "qty_people": 1, "qty_days": 1, "unit": "项"})
n1 = len(r["quote"]["items"]); sd2 = find(r["quote"]["items"], "场地租赁")
check("B1 新增报价项", n1 == n0 + 1 and sd2 is not None, f"{n0}→{n1}")
r = call("DELETE", f"/api/projects/p1/quote/items/{sd2['id']}")
check("B1 删除报价项", len(r["quote"]["items"]) == n0, f"→{len(r['quote']['items'])}")
# 自洽：实收=明细之和
q = call("GET", "/api/projects/p1/quote")
s = round(sum(i["client_amount"] for i in q["items"]))
check("B6 实收 = 各明细客户金额之和", s == round(q["totals"]["client_price"]), f"和{s} 实收{round(q['totals']['client_price'])}")

# ---------- B5 反推 ----------
log("\n【B5】目标反推（尊重锁定）")
call("POST", "/api/projects/p1/generate", {})
call("PUT", "/api/projects/p1/quote/target", {"target_client_price": 130000})
q = call("GET", "/api/projects/p1/quote")
check("按目标总价13万反推", abs(q["totals"]["client_price"] - 130000) <= 50, f"实收{q['totals']['client_price']}")
call("POST", "/api/projects/p1/generate", {})
q = call("GET", "/api/projects/p1/quote"); dao = find(q["items"], "导演")
call("PUT", f"/api/projects/p1/quote/items/{dao['id']}", {"is_locked": True})
c0 = find(call("GET", "/api/projects/p1/quote")["items"], "导演")["client_unit_price"]
call("PUT", "/api/projects/p1/quote/target", {"target_client_price": 200000})
q = call("GET", "/api/projects/p1/quote"); dao = find(q["items"], "导演")
check("反推到20万时锁定的导演不变", dao["client_unit_price"] == c0, f"导演{dao['client_unit_price']}(原{c0})")
check("总价仍达标≈20万", abs(q["totals"]["client_price"] - 200000) <= 100, f"实收{q['totals']['client_price']}")
call("POST", "/api/projects/p1/generate", {})
call("PUT", "/api/projects/p1/quote/target", {"target_margin": 0.35})
q = call("GET", "/api/projects/p1/quote")
check("按目标毛利率35%反推", abs(q["totals"]["gross_margin"] - 0.35) <= 0.01, f"毛利率{round(q['totals']['gross_margin'],3)}")

# ---------- C 导出 ----------
log("\n【C】Excel 导出")
import openpyxl
call("POST", "/api/projects/p1/generate", {})
for ver, tag in [("client", "客户版"), ("internal", "内部版")]:
    data = raw(f"/api/projects/p1/quote.xlsx?version={ver}")
    ws = openpyxl.load_workbook(io.BytesIO(data)).active
    check(f"报价Excel {tag} 可打开", ws["A1"].value and tag in ws["A1"].value, ws["A1"].value)
data = raw("/api/projects/p1/schedule.xlsx")
ws = openpyxl.load_workbook(io.BytesIO(data)).active
hdr = [ws.cell(row=4, column=c).value for c in range(1, 14)]
need = ["项目阶段", "任务名称", "任务说明", "负责人", "协作人", "开始时间", "结束时间", "持续天数", "交付物", "当前状态", "优先级", "风险等级", "备注"]
check("排期Excel 13字段齐全", hdr == need, str(len(hdr)) + "列")

# ---------- D 执行端任务 ----------
log("\n【D】执行端个人任务台")
call("POST", "/api/projects/p1/generate", {})
tasks = call("GET", "/api/tasks?assignee=employee")["tasks"]
p1t = [t for t in tasks if t["project_id"] == "p1"]
check("生成项目→自动派生执行任务", len(p1t) >= 10, f"{len(p1t)}个任务")
t0 = p1t[0]
check("任务含交付物/截止/优先级", bool(t0["deliverable"] and t0["deadline"] and t0["priority"]))
call("POST", f"/api/tasks/{t0['id']}/submit", {"note": "初版完成请查收"})
t = next(x for x in call("GET", "/api/tasks?assignee=employee")["tasks"] if x["id"] == t0["id"])
check("提交成果→状态submitted", t["status"] == "submitted")
msgs = call("GET", "/api/messages/p1?session_id=default&role=boss")["messages"]
check("提交后老板端收到消息", any("张导提交" in m["content"] for m in msgs if m["role"] == "ai"))
call("POST", f"/api/tasks/{t0['id']}/feedback", {"note": "客户素材未到可能延期"})
check("反馈→老板端收到", any("反馈" in m["content"] for m in call("GET", "/api/messages/p1?session_id=default&role=boss")["messages"] if m["role"] == "ai"))

# ---------- 权限隔离（核心：员工拿不到任何预算金额）----------
log("\n【权限】员工问预算不泄露金额（即便老板刚聊过预算）")
chat("现在毛利率多少和实收多少", "boss")          # 老板先聊一轮预算，制造历史
import time as _t; _t.sleep(3)
FIG = ["130023", "148599", "118860", "124860", "141429", "166404", "90270", "60180"]
r2 = chat("现在预算多少给客户报价多少", "employee")
leaked = any(f in r2.replace(",", "").replace("，", "") for f in FIG)
check("员工查预算→不泄露任何金额", not leaked, r2[:40])
_t.sleep(3)
r3 = chat("我负责的任务到哪一步了", "employee")
check("员工可正常对话(非报错)", "【系统】" not in r3 and len(r3) > 0, r3[:30])

# ---------- E 团队 ----------
log("\n【E】团队按阶段动态 + 拉群")
call("POST", "/api/projects/p1/generate", {})
team = call("GET", "/api/projects/p1/team")
check("默认团队已配置(含PM)", len(team["members"]) >= 5 and any(m["is_pm"] for m in team["members"]), f"{len(team['members'])}人")
check("成员按阶段分布", len(set(m["stage"] for m in team["members"])) >= 3)
r = call("POST", "/api/projects/p1/team", {"name": "小测", "role": "摄影助理", "stage": "拍摄"})
w = next(m for m in r["team"]["members"] if m["name"] == "小测")
r = call("PUT", f"/api/projects/p1/team/{w['id']}", {"is_pm": True})
check("设为PM后全项目唯一PM", sum(1 for m in r["team"]["members"] if m["is_pm"]) == 1)
r = call("DELETE", f"/api/projects/p1/team/{w['id']}")
check("移出成员", not any(m["name"] == "小测" for m in r["team"]["members"]))
r = call("POST", "/api/projects/p1/groups", {"name": "拍摄执行组", "members": "张导、摄影、制片", "purpose": "对接现场"})
check("PM拉群", any(g["name"] == "拍摄执行组" for g in r["team"]["groups"]))
# clean test group
gid = next(g["id"] for g in r["team"]["groups"] if g["name"] == "拍摄执行组")
call("DELETE", f"/api/projects/p1/groups/{gid}")

# ---------- 文件上传/下载 ----------
log("\n【文件】上传 Brief 可下载（修无原件）")
u = upload("/api/projects/p1/brief/upload", "测试Brief.txt", "测试内容".encode())
a = next((x for x in call("GET", "/api/projects/p1/assets")["assets"] if x["name"] == "测试Brief.txt"), None)
check("上传Brief后资产可下载", a and a["downloadable"], "downloadable=%s" % (a["downloadable"] if a else "N/A"))

# ---------- 收尾 ----------
call("POST", "/api/projects/p1/generate", {})  # 重置 p1 干净
log("\n" + "="*72)
log(f"测试结果：通过 {P} / 共 {P+F}    " + ("✅ 全部通过" if F == 0 else f"❌ {F} 项未过"))
log("="*72)

open("/Users/bytedance/Documents/Github/AI_Producer/_test_output.txt", "w", encoding="utf-8").write("\n".join(LINES))
